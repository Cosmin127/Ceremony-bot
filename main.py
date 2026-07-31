from pathlib import Path
from datetime import datetime
import json
import sys

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from parser import parse_document
from roblox import lookup


def autosize(ws):

    for column in ws.columns:

        length = 0

        column_letter = get_column_letter(column[0].column)

        for cell in column:

            if cell.value is None:
                continue

            length = max(length, len(str(cell.value)))

        ws.column_dimensions[column_letter].width = min(length + 3, 60)

def process_file(
    input_file: Path,
    output="excel",
    reason="",
    logged_by="",
    exclude_clasps=None,
):

    print(f"Reading {input_file.name}")

    text = input_file.read_text(
        encoding="utf-8"
    )

    rows, skipped = parse_document(
        text,
        exclude_clasps or {},
    )

    print("\nResolving Roblox profiles...\n")

    failed = 0

    for row in rows:

        profile = lookup(row["username"])

        if profile is None:

            print(f"Couldn't find profile for {row['username']}")

            profile = ""

            failed += 1

        row["profile"] = profile

    output_dir = Path("output")
    output_dir.mkdir(exist_ok=True)

    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")

    today = datetime.now().strftime("%d/%m/%y")

    if output == "json":

        output_file = output_dir / f"medals_{timestamp}.json"

        json_rows = []

        for row in rows:

            json_rows.append({
            
                "username": row["username"],
                "medal": row["medal"],
                "clasp": row["clasp"],
            
            })

        with open(output_file, "w", encoding="utf-8") as f:

            json.dump(
                json_rows,
                f,
                indent=4,
                ensure_ascii=False
            )

    else:

        wb = Workbook()

        ws = wb.active

        ws.title = "Medals"

        headers = [

            "Name",
            "Korps",
            "Regiment",
            "Roblox Profile",
            "Medal",
            "Medal Clasp",
            "Reason",
            "Date",
            "Logged By",

        ]

        ws.append(headers)

        for cell in ws[1]:

            cell.font = Font(bold=True)

        for row in rows:

            ws.append([
            
                row["username"],
                "",
                row["regiment"],
                row["profile"],
                row["medal"],
                row["clasp"],
                reason,
                today,
                logged_by,
            
            ])

            profile_cell = ws.cell(
                row=ws.max_row,
                column=4
            )

            if row["profile"]:

                profile_cell.hyperlink = row["profile"]
                profile_cell.style = "Hyperlink"

        ws.freeze_panes = "A2"

        autosize(ws)

        output_file = output_dir / f"medals_{timestamp}.xlsx"

        wb.save(output_file)

    return output_file


def main():

    if len(sys.argv) != 2:

        print("Usage:")
        print("python main.py ceremony.txt")
        return

    input_file = Path(sys.argv[1])

    if not input_file.exists():

        print("Input file not found.")
        return

    output = process_file(input_file)

    print()

    print("Finished.")
    print(f"Output saved to:\n{output}")


if __name__ == "__main__":

    main()
